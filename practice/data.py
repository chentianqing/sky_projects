# -*- coding: utf-8 -*-

# @Project:   sky_projects
# @Time:      2021-02-03 15:09
# @Author  :    TianQing Chen
# @File    :    data.py
# @function :   please note~
import openpyxl

# with open("/Users/chentianqing/Documents/fusion/work/内部错误码数据.xls","r+") as file:
#     print(file)

wb = openpyxl.load_workbook("/Users/chentianqing/Documents/fusion/work/内部错误码数据.xlsx")

#获取excel 两列row1,row2的数据列表。
def ErrorCode(row1,row2,sheet):
    sh = wb[sheet]
    maxrow = sh.max_row  # 最大行
    maxcol = sh.max_column  # 最大列
    all_dist = {}
    for row in range(1,maxrow+1):
        code = sh.cell(row,row1)
        level = sh.cell(row,row2)
        all_dist.update({code.value : level.value})
        # print(code.value,level.value)
    return all_dist

if __name__ == '__main__':
    all = ErrorCode(3,5,"内部错误码列表")
    core = ErrorCode(1,4,"核心错误码")
    for key,value in core.items():
        if key in all.keys():
            print(key,value)
